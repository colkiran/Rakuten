
from openpyxl import load_workbook
from openpyxl.chart import Reference, BarChart

wb = load_workbook("FirstExcel.xlsx")

wb.create_sheet("Chart")

wb.active = wb['Chart']

ws = wb.active

data = [
    ('Products', 'Sales'),
    ('Pepsi', 450),
    ('Coke', 360),
    ('Sprite', 275),
    ('Mirinda', 310),
    ('Thumbs up', 485),
    ('Fanta', 230),
    ('Slice', 250)
]

print(f"data :{data}")

for row in data:
    ws.append(row)

dataRef = Reference(ws, min_row= 2, min_col=2, max_row=8)
labelRef = Reference(ws, min_row=2, min_col=1, max_row=8)

chart = BarChart()
chart.add_data(dataRef)
chart.set_categories(labelRef)
chart.x_axis.title = "Products"
chart.y_axis.title = "Sales in million"
chart.title = "Beverage Sales"

ws.add_chart(chart, "E9")
wb.save("FirstExcel.xlsx")


