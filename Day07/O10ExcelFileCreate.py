
from openpyxl import Workbook
from datetime import datetime

wb = Workbook()

ws = wb.active

ws.title = "Rakuten_Sheet1"

ws['C5'] = "Hello World"
ws['D5'] = 4500
ws['E5'] = datetime.now()

wb.save("FirstExcel.xlsx")
